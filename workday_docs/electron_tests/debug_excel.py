import openpyxl
from pathlib import Path

excel_path = Path(__file__).parent / 'WD_Test_Scenarios_Master.xlsx'
print(f"Reading: {excel_path}")
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Print headers
headers = [cell.value for cell in ws[1]]
print("Headers:")
for i, h in enumerate(headers):
    print(f"  {i}: '{h}'")

# Print first few rows for Peakon
print("\nFirst Peakon scenario:")
for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
    row_dict = dict(zip(headers, row))
    if row_dict.get('Functional Area') == 'Peakon':
        print(f"  Row data: {row_dict}")
        break
