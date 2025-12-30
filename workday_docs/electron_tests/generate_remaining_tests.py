#!/usr/bin/env python3
"""
Generate Electron tests for remaining small functional areas.
"""

import openpyxl
import json
import os
from pathlib import Path
from collections import defaultdict

# Area to folder mapping
AREA_FOLDERS = {
    "Peakon": "Peakon",
    "Labor Optimization": "Labor_Optimization",
    "Prism": "Prism",
    "Journey Paths": "Journey_Paths",
    "PLN - Data Management (HCM)": "Data_Management_HCM",
    "PLN - Data Management (FIN)": "Data_Management_FIN",
    "Messaging": "Messaging",
    "Mobile": "Mobile",
    "Candidate Engagement": "Candidate_Engagement",
    "User Experience": "User_Experience"
}

# WSDL mapping based on RAG queries
AREA_WSDL_MAP = {
    "Peakon": ["Human_Resources", "Performance_Management"],
    "Labor Optimization": ["Scheduling", "Time_Tracking", "Staffing"],
    "Prism": ["Prism_Analytics"],
    "Journey Paths": ["Talent", "Learning", "Performance_Management"],
    "PLN - Data Management (HCM)": ["Human_Resources", "Integrations"],
    "PLN - Data Management (FIN)": ["Financial_Management", "Integrations"],
    "Messaging": ["Notification"],
    "Mobile": ["Human_Resources"],
    "Candidate Engagement": ["Recruiting", "Talent"],
    "User Experience": ["Human_Resources"]
}

def generate_test_file(scenario_id, test_scenario, task, step, functional_area):
    """Generate individual test file content."""

    # Determine WSDL services for this area
    wsdls = AREA_WSDL_MAP.get(functional_area, ["Human_Resources"])

    # Determine confidence based on task/step availability
    if not task or task == "N/A" or not step or step == "N/A":
        confidence = "MANUAL"
        automation_note = "No task/step defined - manual testing required"
        test_type = "MANUAL"
    else:
        confidence = "LOW"
        automation_note = "Limited API coverage - verification needed"
        test_type = "API"

    # Sanitize names for IDs
    safe_scenario = scenario_id.replace("-", "_").replace(" ", "_")

    content = f'''/**
 * Test: {test_scenario}
 * Scenario ID: {scenario_id}
 * Functional Area: {functional_area}
 *
 * Confidence: {confidence}
 * Type: {test_type}
 * WSDL Services: {", ".join(wsdls)}
 *
 * NOTE: {automation_note}
 */

const {{ test, expect }} = require('@playwright/test');
const WorkdayAPI = require('../../lib/workday_api');

test.describe('{functional_area} - {scenario_id}', () => {{
  let api;

  test.beforeAll(async () => {{
    api = new WorkdayAPI({{
      tenant: process.env.WORKDAY_TENANT,
      username: process.env.WORKDAY_USERNAME,
      password: process.env.WORKDAY_PASSWORD
    }});
  }});

  test('{test_scenario}', async () => {{
    test.skip({str(confidence == 'MANUAL').lower()}, 'Manual testing required - no task/step defined');

    // TEST SETUP
    const testData = {{
      scenarioId: '{scenario_id}',
      timestamp: new Date().toISOString()
    }};

'''

    if confidence == "MANUAL":
        content += f'''    // MANUAL TEST STEPS:
    // Task: {task if task else "Not specified"}
    // Step: {step if step else "Not specified"}
    //
    // This scenario requires manual verification through Workday UI
    // Recommended approach:
    // 1. Navigate to {functional_area} module
    // 2. Execute: {task if task else "scenario steps"}
    // 3. Verify: {step if step else "expected outcomes"}

    throw new Error('Manual test - implement UI automation or use Workday web services');
'''
    else:
        content += f'''    // TASK: {task}
    // STEP: {step}

    try {{
      // TODO: Implement specific API calls for {functional_area}
      // Recommended WSDL services: {", ".join(wsdls)}

      // Example structure (adapt to specific scenario):
      // const response = await api.soapRequest('{wsdls[0]}', 'Get_*', {{
      //   version: 'v42.0',
      //   ...params
      // }});

      // Verify response
      // expect(response).toBeDefined();

      console.log('Test placeholder - implement API calls');

    }} catch (error) {{
      console.error('Test execution failed:', error.message);
      throw error;
    }}
'''

    content += f'''  }});

  test.afterAll(async () => {{
    // Cleanup if needed
  }});
}});

/**
 * IMPLEMENTATION NOTES:
 *
 * Functional Area: {functional_area}
 * Confidence Level: {confidence}
 *
 * Recommended WSDL Operations:
 * {chr(10).join([f" * - {wsdl}.wsdl - Check operations with workday_rag.py --wsdl {wsdl}" for wsdl in wsdls])}
 *
 * Query for specific operations:
 * python workday_rag.py --wsdl {wsdls[0]}
 *
 * Task Details:
 * {task if task else "No task specified - requires manual definition"}
 *
 * Step Details:
 * {step if step else "No step specified - requires manual definition"}
 */
'''

    return content

def main():
    # Paths
    excel_path = Path(__file__).parent / 'WD_Test_Scenarios_Master.xlsx'
    output_base = Path(__file__).parent

    print(f"Reading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Extract scenarios
    scenarios_by_area = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        functional_area = row_dict.get('Functional Area')

        if functional_area in AREA_FOLDERS:
            scenarios_by_area[functional_area].append({
                'Scenario_ID': row_dict.get('Scenario ID'),
                'Scenario_Name': row_dict.get('Scenario Name'),
                'Task_Step': row_dict.get('Task / Step'),
                'Sub_Task': row_dict.get('Sub Task')
            })

    # Generate tests for each area
    total_generated = 0

    for functional_area, scenarios in scenarios_by_area.items():
        folder_name = AREA_FOLDERS[functional_area]
        output_dir = output_base / folder_name
        output_dir.mkdir(exist_ok=True)

        print(f"\n{functional_area} ({len(scenarios)} scenarios)")
        print(f"Output: {output_dir}")

        area_count = 0
        for scenario in scenarios:
            scenario_id = scenario['Scenario_ID']
            test_scenario = scenario['Scenario_Name']
            task = scenario['Task_Step']
            step = scenario['Sub_Task']

            if not scenario_id:
                print(f"  [SKIP] No Scenario_ID")
                continue

            # Generate filename
            safe_id = scenario_id.replace(" ", "_").replace("-", "_")
            filename = f"{safe_id}.spec.js"
            filepath = output_dir / filename

            # Generate content
            content = generate_test_file(
                scenario_id, test_scenario, task, step, functional_area
            )

            # Write file
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(content)

            total_generated += 1
            area_count += 1
            print(f"  [OK] {filename}")

        print(f"  Generated {area_count} files for {functional_area}")

    print(f"\n[SUCCESS] Total tests generated: {total_generated}")
    print(f"\nGenerated folders:")
    for area, folder in AREA_FOLDERS.items():
        count = len(scenarios_by_area.get(area, []))
        if count > 0:
            print(f"  {folder}/ - {count} tests")

if __name__ == '__main__':
    main()
